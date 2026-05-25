from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import UTC, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from ozon_frame_common import frame_code, latest_output_dir, load_csv_rows, parse_dt


WAREHOUSE_UFA = "Уфа ФБС"
WAREHOUSE_MSK = "МСК Наш Склад"
MAIN_WINDOW_DAYS = 365
RECENT_WINDOW_DAYS = 28
MIN_STOCK_PER_WAREHOUSE = 3


def calculate_article_recommendation(sales_365: int, stock: int) -> int:
    monthly_need_365 = 0
    if sales_365 > 0:
        monthly_need_365 = max(1, (sales_365 * 30 + MAIN_WINDOW_DAYS - 1) // MAIN_WINDOW_DAYS)
    if stock == 0:
        return MIN_STOCK_PER_WAREHOUSE
    if sales_365 > 0:
        return max(monthly_need_365 - stock, MIN_STOCK_PER_WAREHOUSE - stock, 0)
    return 0


def calculate_frame_recommendation(total_sales_28: int, stock: int, has_non_selling_articles: bool) -> int:
    target_stock = total_sales_28
    if has_non_selling_articles:
        target_stock += MIN_STOCK_PER_WAREHOUSE
    return max(target_stock - stock, 0)


def build_family_rows(
    offers_to_name: dict[str, str],
    sales_detail_rows: list[dict],
    stock_rows: list[dict],
) -> list[dict]:
    latest_sale_dt = max(
        (parse_dt(row.get("in_process_at", "")) for row in sales_detail_rows if row.get("in_process_at")),
        default=datetime.now(UTC),
    )
    recent_cutoff_dt = latest_sale_dt - timedelta(days=RECENT_WINDOW_DAYS)
    main_cutoff_dt = latest_sale_dt - timedelta(days=MAIN_WINDOW_DAYS)

    recent_sales: dict[tuple[str, str], int] = defaultdict(int)
    main_sales: dict[tuple[str, str], int] = defaultdict(int)
    for row in sales_detail_rows:
        sale_dt = parse_dt(row.get("in_process_at", ""))
        if sale_dt is None:
            continue
        offer_id = str(row.get("offer_id", ""))
        warehouse_name = str(row.get("warehouse_name", ""))
        qty = int(row.get("quantity", 0) or 0)
        if sale_dt >= recent_cutoff_dt:
            recent_sales[(offer_id, warehouse_name)] += qty
        if sale_dt >= main_cutoff_dt:
            main_sales[(offer_id, warehouse_name)] += qty

    stock_map = {(str(row["offer_id"]), str(row["warehouse_name"])): int(row.get("free_stock", 0) or 0) for row in stock_rows}

    rows: list[dict] = []
    for offer_id in sorted(offers_to_name):
        sales_28_ufa = recent_sales[(offer_id, WAREHOUSE_UFA)]
        sales_365_ufa = main_sales[(offer_id, WAREHOUSE_UFA)]
        stock_ufa = stock_map.get((offer_id, WAREHOUSE_UFA), 0)
        sales_28_msk = recent_sales[(offer_id, WAREHOUSE_MSK)]
        sales_365_msk = main_sales[(offer_id, WAREHOUSE_MSK)]
        stock_msk = stock_map.get((offer_id, WAREHOUSE_MSK), 0)

        rows.append(
            {
                "Артикул": offer_id,
                "Продажи за 28 дней Уфа ФБС": sales_28_ufa,
                "Продажи за 365 дней Уфа ФБС": sales_365_ufa,
                "Остаток Уфа ФБС": stock_ufa,
                "Рекомендация Уфа ФБС": calculate_article_recommendation(sales_365_ufa, stock_ufa),
                "Продажи за 28 дней МСК Наш Склад": sales_28_msk,
                "Продажи за 365 дней МСК Наш Склад": sales_365_msk,
                "Остаток МСК Наш Склад": stock_msk,
                "Рекомендация МСК Наш Склад": calculate_article_recommendation(sales_365_msk, stock_msk),
                "Продажи за 365 дней Итого": sales_365_ufa + sales_365_msk,
                "Наименование": offers_to_name[offer_id],
            }
        )
    rows.sort(
        key=lambda row: (
            -(row["Рекомендация Уфа ФБС"] + row["Рекомендация МСК Наш Склад"]),
            -row["Продажи за 365 дней Итого"],
            row["Артикул"],
        )
    )
    return rows


def aggregate_metrics_for_rows(rows: list[dict], family_prefixes: list[str]) -> dict[str, dict]:
    aggregated: dict[str, dict] = {}
    for source_row in rows:
        key = frame_code(str(source_row["Артикул"]), family_prefixes)
        row = aggregated.setdefault(
            key,
            {
                "Артикул": key,
                "Продажи за 28 дней Уфа ФБС": 0,
                "Продажи за 365 дней Уфа ФБС": 0,
                "Остаток Уфа ФБС": 0,
                "Рекомендация Уфа ФБС": 0,
                "Продажи за 28 дней МСК Наш Склад": 0,
                "Продажи за 365 дней МСК Наш Склад": 0,
                "Остаток МСК Наш Склад": 0,
                "Рекомендация МСК Наш Склад": 0,
                "Продажи за 365 дней Итого": 0,
                "Наименование": "",
                "_ufa_total_articles": 0,
                "_ufa_selling_articles": 0,
                "_msk_total_articles": 0,
                "_msk_selling_articles": 0,
            },
        )

        sales_28_ufa = int(source_row["Продажи за 28 дней Уфа ФБС"])
        sales_365_ufa = int(source_row["Продажи за 365 дней Уфа ФБС"])
        stock_ufa = int(source_row["Остаток Уфа ФБС"])
        sales_28_msk = int(source_row["Продажи за 28 дней МСК Наш Склад"])
        sales_365_msk = int(source_row["Продажи за 365 дней МСК Наш Склад"])
        stock_msk = int(source_row["Остаток МСК Наш Склад"])

        row["Продажи за 28 дней Уфа ФБС"] += sales_28_ufa
        row["Продажи за 365 дней Уфа ФБС"] += sales_365_ufa
        row["Продажи за 28 дней МСК Наш Склад"] += sales_28_msk
        row["Продажи за 365 дней МСК Наш Склад"] += sales_365_msk
        row["Продажи за 365 дней Итого"] += sales_365_ufa + sales_365_msk
        row["Остаток Уфа ФБС"] = max(int(row["Остаток Уфа ФБС"]), stock_ufa)
        row["Остаток МСК Наш Склад"] = max(int(row["Остаток МСК Наш Склад"]), stock_msk)
        row["_ufa_total_articles"] += 1
        row["_msk_total_articles"] += 1
        if sales_28_ufa > 0:
            row["_ufa_selling_articles"] += 1
        if sales_28_msk > 0:
            row["_msk_selling_articles"] += 1
        if not row["Наименование"]:
            row["Наименование"] = str(source_row.get("Наименование", ""))

    for row in aggregated.values():
        row["Рекомендация Уфа ФБС"] = calculate_frame_recommendation(
            int(row["Продажи за 28 дней Уфа ФБС"]),
            int(row["Остаток Уфа ФБС"]),
            int(row["_ufa_selling_articles"]) < int(row["_ufa_total_articles"]),
        )
        row["Рекомендация МСК Наш Склад"] = calculate_frame_recommendation(
            int(row["Продажи за 28 дней МСК Наш Склад"]),
            int(row["Остаток МСК Наш Склад"]),
            int(row["_msk_selling_articles"]) < int(row["_msk_total_articles"]),
        )
        for helper_key in [
            "_ufa_total_articles",
            "_ufa_selling_articles",
            "_msk_total_articles",
            "_msk_selling_articles",
        ]:
            del row[helper_key]

    return aggregated


def write_sheet(ws, rows: list[dict]) -> None:
    headers = [
        "Артикул",
        "Продажи за 28 дней Уфа ФБС",
        "Продажи за 365 дней Уфа ФБС",
        "Остаток Уфа ФБС",
        "Рекомендация Уфа ФБС",
        "Продажи за 28 дней МСК Наш Склад",
        "Продажи за 365 дней МСК Наш Склад",
        "Остаток МСК Наш Склад",
        "Рекомендация МСК Наш Склад",
        "Продажи за 365 дней Итого",
        "Наименование",
    ]
    ws.append(headers)
    numeric_headers = set(headers) - {"Артикул", "Наименование"}

    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    red_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")

    for row in rows:
        ws.append([int(row[h]) if h in numeric_headers else row[h] for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row_idx in range(2, ws.max_row + 1):
        for start_col in (2, 6):
            sales_28 = ws.cell(row=row_idx, column=start_col).value or 0
            sales_365 = ws.cell(row=row_idx, column=start_col + 1).value or 0
            stock = ws.cell(row=row_idx, column=start_col + 2).value or 0
            fill = None
            if sales_28 > 0 and stock == 0:
                fill = red_fill
            elif sales_365 > 0 and stock == 0:
                fill = yellow_fill
            if fill is not None:
                for col_idx in range(start_col, start_col + 4):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    for idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for cell in ws[get_column_letter(idx)]:
            max_len = max(max_len, len("" if cell.value is None else str(cell.value)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 32)


def load_family_inputs(output_root: Path, cabinet: str, prefix: str) -> tuple[list[dict], list[dict], dict[str, str]]:
    family_dir = latest_output_dir(output_root, cabinet, prefix)
    sales_detail = load_csv_rows(family_dir / "sales_detailed.csv")
    current_stock = load_csv_rows(family_dir / "current_fbs_stocks.csv")
    offers_to_name: dict[str, str] = {}
    sales_summary_path = family_dir / "sales_by_offer_and_warehouse.csv"
    if sales_summary_path.exists():
        for row in load_csv_rows(sales_summary_path):
            offers_to_name.setdefault(str(row["offer_id"]), str(row.get("product_name", "")))
    for row in current_stock:
        offers_to_name.setdefault(str(row["offer_id"]), str(row.get("product_name", "")))
    return sales_detail, current_stock, offers_to_name


def parse_family_cabinets(values: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for value in values:
        family, sep, cabinet = value.partition("=")
        family = family.strip()
        cabinet = cabinet.strip()
        if not sep or not family or not cabinet:
            raise ValueError(
                f"Invalid --family-cabinet value {value!r}. Expected format FAMILY_PREFIX=CABINET_NAME."
            )
        mapping[family] = cabinet
    return mapping


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build an Ozon frame replenishment workbook from family exports.")
    parser.add_argument("cabinet", help="Default cabinet used for families without an explicit mapping")
    parser.add_argument("--family", action="append", required=True, help="Family prefix such as K_T_M1X")
    parser.add_argument(
        "--family-cabinet",
        action="append",
        default=[],
        help="Optional explicit source cabinet in the form FAMILY_PREFIX=CABINET_NAME",
    )
    parser.add_argument("--output-root", type=Path, default=Path("outputs"), help="Directory containing family export folders")
    parser.add_argument("--frame-sheet-name", default="рамки", help="Name of the aggregated frame sheet")
    parser.add_argument("--workbook-path", type=Path, help="Target .xlsx path")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    family_cabinets = parse_family_cabinets(args.family_cabinet)
    family_rows_map: dict[str, list[dict]] = {}
    all_rows: list[dict] = []
    for family in args.family:
        source_cabinet = family_cabinets.get(family, args.cabinet)
        sales_detail, current_stock, offers_to_name = load_family_inputs(args.output_root, source_cabinet, family)
        rows = build_family_rows(offers_to_name, sales_detail, current_stock)
        family_rows_map[family] = rows
        all_rows.extend(rows)

    frame_rows = list(aggregate_metrics_for_rows(all_rows, args.family).values())
    frame_rows.sort(
        key=lambda row: (
            -(row["Рекомендация Уфа ФБС"] + row["Рекомендация МСК Наш Склад"]),
            -row["Продажи за 365 дней Итого"],
            row["Артикул"],
        )
    )

    workbook = Workbook()
    first_sheet = True
    for family in args.family:
        ws = workbook.active if first_sheet else workbook.create_sheet(family)
        ws.title = family
        write_sheet(ws, family_rows_map[family])
        first_sheet = False

    frame_ws = workbook.create_sheet(args.frame_sheet_name)
    write_sheet(frame_ws, frame_rows)

    workbook_path = args.workbook_path or (args.output_root / f"{args.cabinet.lower()}_frame_replenishment.xlsx")
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    final_path = workbook_path
    try:
        workbook.save(final_path)
    except PermissionError:
        final_path = workbook_path.with_name(f"{workbook_path.stem}_{datetime.now():%Y%m%d_%H%M%S}{workbook_path.suffix}")
        workbook.save(final_path)
    print(final_path.resolve())


if __name__ == "__main__":
    main()
